import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

def actualizar_pendientes():
    """
    Copia todos los datos de TempFiles.xlsx a Pendientes.xlsx
    en la hoja 'Relatorio Suportes'
    """
    
    # Rutas de archivos
    temp_file = 'TempFiles.xlsx'
    pendientes_file = 'Pendientes.xlsx'
    hoja_destino = 'Relatorio Suportes'
    
    # Validar que existan los archivos
    if not os.path.exists(temp_file):
        print(f'❌ Error: {temp_file} no encontrado')
        return False
    
    if not os.path.exists(pendientes_file):
        print(f'❌ Error: {pendientes_file} no encontrado')
        return False
    
    try:
        print(f'📖 Leyendo {temp_file}...')
        # Leer todos los datos de TempFiles (con encabezados)
        df_temp = pd.read_excel(temp_file, sheet_name=0)
        
        print(f'📊 Datos leídos: {len(df_temp)} filas, {len(df_temp.columns)} columnas')
        print(f'Columnas: {list(df_temp.columns)}')
        
        # Abrir Pendientes.xlsx
        print(f'📖 Abriendo {pendientes_file}...')
        wb = load_workbook(pendientes_file)
        
        # Verificar que existe la hoja
        if hoja_destino not in wb.sheetnames:
            print(f'❌ Error: La hoja "{hoja_destino}" no existe en {pendientes_file}')
            print(f'Hojas disponibles: {wb.sheetnames}')
            return False
        
        ws = wb[hoja_destino]
        
        # Limpiar datos existentes (mantener solo encabezados si existen)
        print(f'🧹 Limpiando datos en "{hoja_destino}"...')
        
        # Encontrar última fila con datos
        max_row = ws.max_row
        if max_row > 1:
            ws.delete_rows(2, max_row)
        
        # Escribir encabezados si no existen
        if ws.max_row == 0:
            for col_num, column_title in enumerate(df_temp.columns, 1):
                ws.cell(row=1, column=col_num, value=column_title)
        
        # Escribir datos
        print(f'✍️ Escribiendo {len(df_temp)} filas de datos...')
        for row_num, row_data in enumerate(df_temp.values, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
        
        # Guardar archivo
        print(f'💾 Guardando {pendientes_file}...')
        wb.save(pendientes_file)
        
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        print(f'✅ Actualización completada: {timestamp}')
        print(f'✅ {len(df_temp)} filas copiadas exitosamente')
        
        return True
        
    except Exception as e:
        print(f'❌ Error durante la actualización: {str(e)}')
        return False

if __name__ == '__main__':
    actualizar_pendientes()